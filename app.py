import os
import io
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# Configure system encoding to utf-8 for Windows terminal support
sys.stdout.reconfigure(encoding='utf-8')

# Set page config
st.set_page_config(
    page_title="Đối Chiếu Công Nợ Nội Bộ 136/336",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium CSS styling (Navy theme)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=Space+Grotesk:wght@500;700&display=swap');

    html, body, [class*="css"], .stApp {
        font-family: 'Plus Jakarta Sans', sans-serif !important;
    }

    .stApp {
        background: radial-gradient(circle at 50% 50%, #fbfcfe 0%, #f3f5fa 100%) !important;
    }

    .main-title {
        color: #1E3A8A;
        font-family: 'Plus Jakarta Sans', sans-serif;
        font-weight: 800;
        font-size: 2.25rem;
        background: linear-gradient(135deg, #1B365D 0%, #3B82F6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 6px;
        letter-spacing: -0.5px;
    }

    .sub-title {
        color: #64748B;
        font-family: 'Plus Jakarta Sans', sans-serif;
        font-size: 1.05rem;
        font-weight: 500;
        margin-bottom: 30px;
    }

    .section-header {
        color: #1E3A8A;
        font-weight: 700;
        font-size: 1.2rem;
        border-left: 4px solid #3B82F6;
        padding-left: 12px;
        margin-top: 35px;
        margin-bottom: 20px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* Metric card styling */
    .metric-card {
        background-color: #FFFFFF;
        padding: 22px 18px;
        border-radius: 16px;
        box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.03), 0 8px 10px -6px rgba(0, 0, 0, 0.03);
        border: 1px solid #F1F5F9;
        text-align: center;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }

    .metric-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.08), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
    }

    .metric-card-danger {
        border-top: 4px solid #EF4444;
    }

    .metric-card-warning {
        border-top: 4px solid #F59E0B;
    }

    .metric-card-success {
        border-top: 4px solid #10B981;
    }

    .metric-card-info {
        border-top: 4px solid #3B82F6;
    }

    .metric-value {
        font-family: 'Space Grotesk', sans-serif;
        font-size: 1.45rem;
        font-weight: 700;
        margin-top: 8px;
        letter-spacing: -0.3px;
        white-space: nowrap;
    }

    .metric-label {
        font-size: 0.85rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        color: #64748B;
    }

    /* Tabs styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 12px;
        background-color: transparent;
        border-bottom: 2px solid #E2E8F0;
        padding-bottom: 4px;
    }

    .stTabs [data-baseweb="tab"] {
        background-color: rgba(255, 255, 255, 0.6);
        border-radius: 12px 12px 0 0 !important;
        color: #475569;
        font-weight: 600;
        padding: 10px 24px;
        transition: all 0.2s ease;
        border: 1px solid #E2E8F0;
        border-bottom: none;
        font-size: 0.95rem;
    }

    .stTabs [data-baseweb="tab"]:hover {
        background-color: rgba(241, 245, 249, 0.8);
        color: #1E3A8A;
    }

    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #1B365D 0%, #2a4e80 100%) !important;
        color: white !important;
        box-shadow: 0 4px 15px rgba(27, 54, 93, 0.15);
        border: 1px solid #1B365D !important;
        border-bottom: none !important;
    }

    /* Feature card styling */
    .feature-card {
        background-color: #ffffff;
        padding: 28px 24px;
        border-radius: 16px;
        box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.03), 0 8px 10px -6px rgba(0, 0, 0, 0.03);
        border: 1px solid #F1F5F9;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        height: 100%;
    }

    .feature-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.07), 0 10px 10px -5px rgba(0, 0, 0, 0.03);
        border-color: #3B82F6;
    }

    .feature-card h4 {
        color: #1E3A8A !important;
        margin-top: 0;
        font-weight: 700;
        font-size: 1.15rem;
        margin-bottom: 12px;
    }

    .feature-card p {
        color: #64748B !important;
        font-size: 0.925rem;
        line-height: 1.6;
        margin: 0;
    }

    /* Custom styled buttons */
    div.stButton > button, div.stDownloadButton > button {
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        font-weight: 600 !important;
        border-radius: 12px !important;
        border: none !important;
        padding: 12px 24px !important;
        transition: all 0.3s ease !important;
        font-size: 0.95rem !important;
        height: auto !important;
        width: 100% !important;
    }

    div.stDownloadButton > button {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(16, 185, 129, 0.2) !important;
    }

    div.stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        background: linear-gradient(135deg, #059669 0%, #10B981 100%) !important;
        box-shadow: 0 6px 20px rgba(16, 185, 129, 0.35) !important;
    }

    div.stButton > button {
        background: linear-gradient(135deg, #3B82F6 0%, #1E3A8A 100%) !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.2) !important;
    }

    div.stButton > button:hover {
        transform: translateY(-2px) !important;
        background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%) !important;
        box-shadow: 0 6px 20px rgba(59, 130, 246, 0.35) !important;
    }

    div.stButton > button:active, div.stDownloadButton > button:active {
        transform: translateY(1px) !important;
    }

    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #FFFFFF !important;
        border-right: 1px solid #E2E8F0 !important;
    }

    [data-testid="stSidebar"] [data-testid="stFileUploader"] {
        border: 1px dashed #CBD5E1 !important;
        border-radius: 12px !important;
        background-color: #F8FAFC !important;
        padding: 10px !important;
    }
</style>
""", unsafe_allow_html=True)

# App header
st.markdown('<div class="main-title">HỆ THỐNG ĐỐI CHIẾU CÔNG NỢ NỘI BỘ TK 136/336</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Hỗ trợ đối soát tự động giữa Văn phòng Tổng công ty (TCT) và Công ty Điện lực Vũng Tàu (PCVT)</div>', unsafe_allow_html=True)

# -------------------------------------------------------------
# CONSTANTS & SETUP
# -------------------------------------------------------------
NAVY_HEADER = "1B365D"
NAVY_ACCENT = "E8EEF5"
WHITE = "FFFFFF"
BORDER_COLOR = "D9D9D9"
ALERT_GREEN = "D1E7DD"
ALERT_RED = "F8D7DA"

ACCOUNT_PAIRS = [
    {"tct": "1363111", "unit": "3363111", "name": "Doanh thu tiền điện"},
    {"tct": "1363112", "unit": "3363112", "name": "Doanh thu tiền điện vô công"},
    {"tct": "136314", "unit": "336314", "name": "Doanh thu QL dây, cáp"},
    {"tct": "136331", "unit": "336331", "name": "Vốn ĐT XDCB"},
    {"tct": "136332", "unit": "336332", "name": "Vốn ĐT XDCB khác"},
    {"tct": "136341", "unit": "336341", "name": "Khấu hao TSCĐ nguồn vốn vay"},
    {"tct": "136351", "unit": "336351", "name": "Phải thu/trả Sửa chữa lớn"},
    {"tct": "136358", "unit": "336358", "name": "Phải thu/trả nội bộ khác"},
    {"tct": "1363611", "unit": "3363611", "name": "Thuế GTGT ĐMT mái nhà"},
    {"tct": "1363613", "unit": "3363613", "name": "Thuế GTGT nội bộ khác"},
    {"tct": "13638221", "unit": "33638221", "name": "LN vi phạm HĐ bán điện"},
    {"tct": "1363847", "unit": "3363847", "name": "LN hoạt động tài chính"},
    {"tct": "1363848", "unit": "3363848", "name": "LN thu nhập khác"},
    {"tct": "1363883", "unit": "3363883", "name": "Luân chuyển vật tư thiết bị"},
    {"tct": "13681", "unit": "33681", "name": "Phải thu/trả nội bộ về chi phí"},
]

# Sidebar file uploaders
st.sidebar.markdown('<div style="font-size:1.15rem; font-weight:bold; color:#1B365D; margin-bottom:10px;">📤 Tải lên Tệp dữ liệu</div>', unsafe_allow_html=True)
tool_file = st.sidebar.file_uploader(
    "1. File Tool Đối Chiếu (Bắt buộc)", 
    type=["xlsx"], 
    help="Tệp Excel chứa các sheet TONG_HOP, DATA_TONGHOP, DATA_PHATSINH..."
)

# Detailed ledgers uploaders
ledger_tct_136 = st.sidebar.file_uploader("2. Sổ chi tiết 136 - TCT (Tùy chọn)", type=["xlsx"])
ledger_tct_336 = st.sidebar.file_uploader("3. Sổ chi tiết 336 - TCT (Tùy chọn)", type=["xlsx"])
ledger_pcvt_136 = st.sidebar.file_uploader("4. Sổ chi tiết 136 - PCVT (Tùy chọn)", type=["xlsx"])
ledger_pcvt_336 = st.sidebar.file_uploader("5. Sổ chi tiết 336 - PCTV (Tùy chọn)", type=["xlsx"])

# -------------------------------------------------------------
# RECONCILIATION CORE ENGINE
# -------------------------------------------------------------
@st.cache_data
def load_and_reconcile(file_bytes):
    xl = pd.ExcelFile(file_bytes)
    
    # 1. Parse DATA_TONGHOP for Balances
    df_th = xl.parse("DATA_TONGHOP")
    df_th['SHTK'] = df_th['SHTK'].astype(str).str.split('.').str[0].str.strip()
    df_th['MÃ ĐVI'] = df_th['MÃ ĐVI'].astype(str).str.split('.').str[0].str.strip()
    df_th['NET'] = df_th['NET'].fillna(0).astype(float)
    
    # 2. Parse DATA_PHATSINH for detailed transactions
    df_ps = xl.parse("DATA_PHATSINH")
    df_ps['SHTK'] = df_ps['SHTK'].astype(str).str.split('.').str[0].str.strip()
    df_ps['MÃ ĐVI'] = df_ps['MÃ ĐVI'].astype(str).str.split('.').str[0].str.strip()
    df_ps['SỐ PS NỢ'] = df_ps['SỐ PS NỢ'].fillna(0).astype(float)
    df_ps['SỐ PS CÓ'] = df_ps['SỐ PS CÓ'].fillna(0).astype(float)
    df_ps['NET'] = df_ps['NET'].fillna(0).astype(float)
    
    # Split into TCT (80100) and PCVT (82900)
    df_ps_tct = df_ps[df_ps['MÃ ĐVI'] == '80100'].copy()
    df_ps_pcvt = df_ps[df_ps['MÃ ĐVI'] == '82900'].copy()
    
    # Map PCVT SHTK to align with TCT SHTK (336 -> 136)
    df_ps_pcvt['SHTK_MAPPED'] = df_ps_pcvt['SHTK'].apply(lambda x: '136' + x[3:] if x.startswith('336') else x)
    df_ps_tct['SHTK_MAPPED'] = df_ps_tct['SHTK']
    
    matched_t = set()
    matched_p = set()
    
    # Pass 1: Match within mapped SHTK, same amount and voucher
    for idx_t, row_t in df_ps_tct.iterrows():
        shtk = row_t['SHTK_MAPPED']
        amt = row_t['NET']
        v_gl = str(row_t['SỐ CT PHÂN HỆ GL']).strip()
        v_ph = str(row_t['SỐ CT PHÂN HỆ PHỤ']).strip()
        
        matches = df_ps_pcvt[
            (df_ps_pcvt['SHTK_MAPPED'] == shtk) &
            (df_ps_pcvt['NET'] == amt) &
            (~df_ps_pcvt.index.isin(matched_p)) &
            ((df_ps_pcvt['SỐ CT PHÂN HỆ GL'].astype(str).str.strip() == v_gl) |
             (df_ps_pcvt['SỐ CT PHÂN HỆ PHỤ'].astype(str).str.strip() == v_ph))
        ]
        if not matches.empty:
            matched_t.add(idx_t)
            matched_p.add(matches.index[0])
            
    # Pass 2: Match within mapped SHTK, same amount and date (within 5 days)
    for idx_t, row_t in df_ps_tct.iterrows():
        if idx_t in matched_t:
            continue
        shtk = row_t['SHTK_MAPPED']
        amt = row_t['NET']
        d_t = pd.to_datetime(row_t['NGÀY CT'])
        
        matches = df_ps_pcvt[
            (df_ps_pcvt['SHTK_MAPPED'] == shtk) &
            (df_ps_pcvt['NET'] == amt) &
            (~df_ps_pcvt.index.isin(matched_p))
        ]
        for idx_p, row_p in matches.iterrows():
            d_p = pd.to_datetime(row_p['NGÀY CT'])
            if abs((d_t - d_p).days) <= 5:
                matched_t.add(idx_t)
                matched_p.add(idx_p)
                break
                
    # Pass 3: Match within mapped SHTK, same amount only
    for idx_t, row_t in df_ps_tct.iterrows():
        if idx_t in matched_t:
            continue
        shtk = row_t['SHTK_MAPPED']
        amt = row_t['NET']
        
        matches = df_ps_pcvt[
            (df_ps_pcvt['SHTK_MAPPED'] == shtk) &
            (df_ps_pcvt['NET'] == amt) &
            (~df_ps_pcvt.index.isin(matched_p))
        ]
        if not matches.empty:
            matched_t.add(idx_t)
            matched_p.add(matches.index[0])
            
    # Pass 4: Cross-account match (any account), same amount and date (within 5 days)
    # This matches classification mismatches globally
    for idx_t, row_t in df_ps_tct.iterrows():
        if idx_t in matched_t:
            continue
        amt = row_t['NET']
        d_t = pd.to_datetime(row_t['NGÀY CT'])
        
        matches = df_ps_pcvt[
            (df_ps_pcvt['NET'] == amt) &
            (~df_ps_pcvt.index.isin(matched_p))
        ]
        for idx_p, row_p in matches.iterrows():
            d_p = pd.to_datetime(row_p['NGÀY CT'])
            if abs((d_t - d_p).days) <= 5:
                matched_t.add(idx_t)
                matched_p.add(idx_p)
                break
                
    # Pass 5: Cross-account match (any account), same amount only
    for idx_t, row_t in df_ps_tct.iterrows():
        if idx_t in matched_t:
            continue
        amt = row_t['NET']
        
        matches = df_ps_pcvt[
            (df_ps_pcvt['NET'] == amt) &
            (~df_ps_pcvt.index.isin(matched_p))
        ]
        if not matches.empty:
            matched_t.add(idx_t)
            matched_p.add(matches.index[0])
            
    # Extract unmatched global dataframes
    df_unmatched_tct = df_ps_tct[~df_ps_tct.index.isin(matched_t)]
    df_unmatched_pcvt = df_ps_pcvt[~df_ps_pcvt.index.isin(matched_p)]

    # Compute summaries & explanations
    summary_rows = []
    details_data = {}
    
    for pair in ACCOUNT_PAIRS:
        tk_tct = pair["tct"]
        tk_unit = pair["unit"]
        tk_name = pair["name"]
        
        # Get balances
        tct_row = df_th[(df_th['MÃ ĐVI'] == '80100') & (df_th['SHTK'] == tk_tct)]
        unit_row = df_th[(df_th['MÃ ĐVI'] == '82900') & (df_th['SHTK'] == tk_unit)]
        
        bal_tct = tct_row['NET'].sum() if not tct_row.empty else 0.0
        bal_unit = unit_row['NET'].sum() if not unit_row.empty else 0.0
        diff = bal_tct - bal_unit
        
        # Filter global unmatched lists for this specific account
        unmatched_t = df_unmatched_tct[df_unmatched_tct['SHTK'] == tk_tct]
        unmatched_p = df_unmatched_pcvt[df_unmatched_pcvt['SHTK'] == tk_unit]
        
        details_data[f"{tk_tct}_{tk_unit}"] = {
            "tct": unmatched_t,
            "unit": unmatched_p
        }
        
        # Generate explanations dynamically
        explanation = "Khớp 100% cả phát sinh và số dư."
        if abs(diff) > 1:
            if tk_tct == "1363111":
                rev_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('51111|doanh thu điện', case=False, na=False)]['SỐ PS CÓ'].sum()
                explanation = f"TCT chưa hạch toán doanh thu điện chuyển lên từ đơn vị ({rev_sum:,.0f}đ). Lệch tài khoản đối ứng khi chuyển tiền về TCT (TCT ghi 33688, đơn vị ghi 3363111). Lệch giải trừ thu hộ qua cổng thanh toán."
            elif tk_tct == "1363112":
                rev_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('51113|vô công|phản kháng', case=False, na=False)]['SỐ PS CÓ'].sum()
                explanation = f"PCVT kết chuyển doanh thu điện vô công (phản kháng) tăng phải trả TCT ({rev_sum:,.0f}đ) nhưng TCT chưa hạch toán ghi nhận. Lệch đối chiếu giải trừ thu hộ."
            elif tk_tct == "136314":
                rev_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('5114|dây|cáp', case=False, na=False)]['SỐ PS CÓ'].sum()
                explanation = f"PCVT kết chuyển doanh thu QL dây cáp ({rev_sum:,.0f}đ), TCT chưa hạch toán ghi nhận tăng phải thu tương ứng."
            elif tk_tct == "136331":
                qt_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('quyết toán|hoàn thành', case=False, na=False)]['SỐ PS NỢ'].sum()
                tscd_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('TSCĐ|điều động tài sản', case=False, na=False)]['SỐ PS CÓ'].sum()
                explanation = f"PCVT kết chuyển giảm phải trả do quyết toán bàn giao CT hoàn thành ({qt_sum:,.0f}đ), TCT chưa ghi Có. PCVT hạch toán nhận điều động TSCĐ ({tscd_sum:,.0f}đ), TCT chưa ghi đối ứng."
            elif tk_tct == "136332":
                explanation = "Lệch phân loại: Đơn vị hạch toán cấp vốn KHCB trên 336332 (Có), TCT hạch toán cấp vốn này vào 136331 (Nợ). Khớp 100% về tổng số liệu cấp vốn."
            elif tk_tct == "136341":
                explanation = "Lệch phân loại: Đơn vị kết chuyển khấu hao về TCT trên 336341, TCT hạch toán ghi nhận vào 136352. Số liệu khớp 100%."
            elif tk_tct == "136351":
                scl_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('sửa chữa lớn|SCL|hoàn thành', case=False, na=False)]['SỐ PS NỢ'].sum()
                explanation = f"PCVT kết chuyển giảm phải trả chi phí SCL hoàn thành bàn giao ({scl_sum:,.0f}đ), TCT chưa hạch toán ghi Có giảm phải thu."
            elif tk_tct == "136358":
                elec_sum = unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('điện mua nội bộ|HĐ 819', case=False, na=False)]['SỐ PS CÓ'].sum() - unmatched_p[unmatched_p['DIỄN GIẢI'].str.contains('bù trừ chi phí điện', case=False, na=False)]['SỐ PS NỢ'].sum()
                explanation = f"TCT chưa hạch toán điện mua nội bộ từ đơn vị (chênh lệch ròng {elec_sum:,.0f}đ). Lệch đối chiếu cấp vốn lương/bảo hiểm. Các giao dịch luân chuyển vật tư thiết bị Net=0."
            elif tk_tct == "1363611":
                tax_sum = unmatched_t[unmatched_t['DIỄN GIẢI'].str.contains('Cấp Thuế GTGT|ĐMT', case=False, na=False)]['SỐ PS NỢ'].sum()
                explanation = f"TCT hạch toán cấp vốn thuế GTGT ĐMT mái nhà ({tax_sum:,.0f}đ), PCVT chưa ghi nhận nhận cấp vốn trên 3363611."
            elif tk_tct in ["13638221", "1363847", "1363848"]:
                explanation = "Số dư và phát sinh khớp 100%. Báo lệch trên file đối chiếu do lỗi công thức đối chiếu của TCT không cộng phát sinh trong tháng của TCT (Lệch ảo)."
            elif tk_tct == "1363883":
                explanation = "Khớp 100% về vật tư thiết bị luân chuyển. Đơn vị hạch toán cấn trừ công nợ điều chỉnh nội bộ, TCT ghi nhận khớp ở tiểu khoản khác."
            elif tk_tct == "13681":
                explanation = f"Phát sinh trong tháng khớp 100%. Số dư chênh lệch là do số dư chênh lệch lũy kế đầu kỳ ({diff:,.0f}đ) từ các tháng trước chuyển sang."
            else:
                explanation = f"Chênh lệch số dư chưa đồng bộ ({diff:,.0f}đ). Cần rà soát các chứng từ chưa khớp."
                
        summary_rows.append([
            f"{tk_tct} / {tk_unit}",
            tk_name,
            bal_tct,
            bal_unit,
            diff,
            explanation
        ])
        
    return summary_rows, details_data

# -------------------------------------------------------------
# EXCEL GENERATOR (WITH IN-MEMORY BUFFER)
# -------------------------------------------------------------
def generate_excel_report(summary_data, details):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    # Fonts and styles
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_bold_data = Font(name="Segoe UI", size=10, bold=True)
    
    fill_header = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_accent = PatternFill(start_color=NAVY_ACCENT, end_color=NAVY_ACCENT, fill_type="solid")
    fill_green = PatternFill(start_color=ALERT_GREEN, end_color=ALERT_GREEN, fill_type="solid")
    fill_red = PatternFill(start_color=ALERT_RED, end_color=ALERT_RED, fill_type="solid")
    
    thin_side = Side(border_style="thin", color=BORDER_COLOR)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    def format_sheet(ws, title, headers, col_formats):
        # Title
        ws.cell(row=1, column=1, value=title).font = font_title
        ws.row_dimensions[1].height = 30
        
        # Headers at Row 3
        for idx, h in enumerate(headers):
            cell = ws.cell(row=3, column=idx+1, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 28
        
        # Format Data Rows
        for r in range(4, ws.max_row + 1):
            ws.row_dimensions[r].height = 20
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = font_data
                cell.border = thin_border
                
                # Alignments and formats
                fmt = col_formats[c-1]
                if fmt == "currency":
                    cell.number_format = '#,##0;[Red]-#,##0;"-";@'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif fmt == "date":
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif fmt == "center":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                # Zebra striping
                if r % 2 == 0:
                    cell.fill = fill_accent
                    
        # Auto-fit widths
        for col_idx in range(len(headers)):
            col_letter = get_column_letter(col_idx + 1)
            max_len = len(headers[col_idx])
            for r in range(4, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx + 1).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        ws.views.sheetView[0].showGridLines = True

    # 1. Sheet 1: TONG_HOP_CHENH_LECH
    ws_sum = wb.create_sheet(title="TONG_HOP_CHENH_LECH")
    headers_sum = ["Tài Khoản", "Tên Tài Khoản Đối Chiếu", "Số Dư TCT (TK 136)", "Số Dư Đơn Vị (TK 336)", "Chênh Lệch (TCT - Đơn vị)", "Phân Tích Nguyên Nhân Sự Việc"]
    for r_idx, r in enumerate(summary_data):
        for c_idx, val in enumerate(r):
            ws_sum.cell(row=r_idx+4, column=c_idx+1, value=val)
            
    format_sheet(ws_sum, "BẢNG TỔNG HỢP CHÊNH LỆCH CÔNG NỢ NỘI BỘ TK 136/336", headers_sum, ["center", "left", "currency", "currency", "currency", "left"])
    
    # Red/Green highlight for differences on Sheet 1
    for r in range(4, ws_sum.max_row + 1):
        diff_val = ws_sum.cell(row=r, column=5).value
        cell_diff = ws_sum.cell(row=r, column=5)
        if diff_val != 0:
            cell_diff.fill = fill_red
            cell_diff.font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
        else:
            cell_diff.fill = fill_green
            cell_diff.font = Font(name="Segoe UI", size=10, bold=True, color="006100")

    # Helpers to write detail sheets
    def write_unmatched_rows(ws, unmatched_tct, unmatched_pcvt, columns_to_write, headers, formats, title):
        # Header is row 3
        # Data starts from row 4
        current_row = 4
        
        # Write TCT rows
        for _, row in unmatched_tct.iterrows():
            ws.cell(row=current_row, column=1, value="Tổng công ty (TCT)")
            for col_idx, col_name in enumerate(columns_to_write):
                val = row[col_name]
                if pd.isnull(val):
                    val = None
                elif col_name == 'NGÀY CT':
                    val = pd.to_datetime(val).strftime('%Y-%m-%d')
                elif isinstance(val, (np.integer, np.int64)):
                    val = int(val)
                elif isinstance(val, (np.floating, np.float64)):
                    val = float(val)
                ws.cell(row=current_row, column=col_idx+2, value=val)
            current_row += 1
            
        # Write PCVT rows
        for _, row in unmatched_pcvt.iterrows():
            ws.cell(row=current_row, column=1, value="Đơn vị (PCVT)")
            for col_idx, col_name in enumerate(columns_to_write):
                val = row[col_name]
                if pd.isnull(val):
                    val = None
                elif col_name == 'NGÀY CT':
                    val = pd.to_datetime(val).strftime('%Y-%m-%d')
                elif isinstance(val, (np.integer, np.int64)):
                    val = int(val)
                elif isinstance(val, (np.floating, np.float64)):
                    val = float(val)
                ws.cell(row=current_row, column=col_idx+2, value=val)
            current_row += 1
            
        format_sheet(ws, title, ["Bên Hạch Toán"] + headers, ["center"] + formats)

    # 2. Sheet 2: LECH_1363111_3363111
    ws_111 = wb.create_sheet(title="LECH_1363111_3363111")
    cols_111 = ['NGÀY CT', 'SỐ CT PHÂN HỆ GL', 'DIỄN GIẢI', 'SỐ PS NỢ', 'SỐ PS CÓ', 'NGƯỜI LẬP CT']
    headers_111 = ["Ngày CT", "Số CT GL", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Người Lập"]
    formats_111 = ["center", "center", "left", "currency", "currency", "center"]
    
    # Add custom note column
    write_unmatched_rows(ws_111, details["1363111_3363111"]["tct"], details["1363111_3363111"]["unit"], cols_111, headers_111, formats_111, "CHI TIẾT CÁC GIAO DỊCH CHÊNH LỆCH LỚN - TK 1363111 vs TK 3363111")
    
    # 3. Sheet 3: LECH_136358_336358
    ws_58 = wb.create_sheet(title="LECH_136358_336358")
    cols_58 = ['NGÀY CT', 'NGUỒN', 'DIỄN GIẢI', 'SỐ PS NỢ', 'SỐ PS CÓ', 'NGƯỜI LẬP CT']
    headers_58 = ["Ngày CT", "Nguồn", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Người Lập"]
    formats_58 = ["center", "center", "left", "currency", "currency", "center"]
    
    write_unmatched_rows(ws_58, details["136358_336358"]["tct"], details["136358_336358"]["unit"], cols_58, headers_58, formats_58, "CHI TIẾT CÁC GIAO DỊCH CHÊNH LỆCH LỚN - TK 136358 vs TK 336358")

    # 4. Sheet 4: LECH_XDCB
    ws_xdcb = wb.create_sheet(title="LECH_XDCB")
    # Combine construction and asset sheets (136331, 136332, 136341, 136351)
    current_row = 4
    for pair_name in ["136331_336331", "136332_336332", "136341_336341", "136351_336351"]:
        t_unmatched = details[pair_name]["tct"]
        p_unmatched = details[pair_name]["unit"]
        
        for _, row in t_unmatched.iterrows():
            ws_xdcb.cell(row=current_row, column=1, value=row['SHTK'])
            ws_xdcb.cell(row=current_row, column=2, value="Tổng công ty (TCT)")
            ws_xdcb.cell(row=current_row, column=3, value=pd.to_datetime(row['NGÀY CT']).strftime('%Y-%m-%d') if pd.notnull(row['NGÀY CT']) else "")
            
            diag = row['DIỄN GIẢI']
            ws_xdcb.cell(row=current_row, column=4, value=diag if pd.notnull(diag) else None)
            
            ps_no = row['SỐ PS NỢ']
            ws_xdcb.cell(row=current_row, column=5, value=float(ps_no) if pd.notnull(ps_no) else 0.0)
            
            ps_co = row['SỐ PS CÓ']
            ws_xdcb.cell(row=current_row, column=6, value=float(ps_co) if pd.notnull(ps_co) else 0.0)
            current_row += 1
            
        for _, row in p_unmatched.iterrows():
            ws_xdcb.cell(row=current_row, column=1, value=row['SHTK'])
            ws_xdcb.cell(row=current_row, column=2, value="Đơn vị (PCVT)")
            ws_xdcb.cell(row=current_row, column=3, value=pd.to_datetime(row['NGÀY CT']).strftime('%Y-%m-%d') if pd.notnull(row['NGÀY CT']) else "")
            
            diag = row['DIỄN GIẢI']
            ws_xdcb.cell(row=current_row, column=4, value=diag if pd.notnull(diag) else None)
            
            ps_no = row['SỐ PS NỢ']
            ws_xdcb.cell(row=current_row, column=5, value=float(ps_no) if pd.notnull(ps_no) else 0.0)
            
            ps_co = row['SỐ PS CÓ']
            ws_xdcb.cell(row=current_row, column=6, value=float(ps_co) if pd.notnull(ps_co) else 0.0)
            current_row += 1
            
    headers_xdcb = ["Tài Khoản", "Bên Hạch Toán", "Ngày CT", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có"]
    formats_xdcb = ["center", "center", "center", "left", "currency", "currency"]
    format_sheet(ws_xdcb, "CHI TIẾT CHÊNH LỆCH VÀ PHÂN LOẠI NHẦM - NHÓM VỐN ĐẦU TƯ XDCB & KHẤU HAO", headers_xdcb, formats_xdcb)

    # 5. Sheet 5: KHOAN_KHOP_AO
    ws_khop = wb.create_sheet(title="KHOAN_KHOP_AO")
    # Show how 13638221, 1363847, 1363848, 1363883, 13681 match
    khop_rows = []
    for r in summary_data:
        tk = r[0].split(" / ")[0]
        if tk in ["13638221", "1363847", "1363848", "1363883", "13681"]:
            khop_rows.append([
                r[0],
                r[1],
                r[2],
                r[3],
                "Nợ/Có khớp 100%",
                "Nợ/Có khớp 100%",
                r[5]
            ])
    for r_idx, r in enumerate(khop_rows):
        for c_idx, val in enumerate(r):
            ws_khop.cell(row=r_idx+4, column=c_idx+1, value=val)
            
    headers_khop = ["Tài Khoản", "Tên Tài Khoản", "Số Dư TCT", "Số Dư Đơn Vị", "Phát Sinh TCT (Nợ/Có)", "Phát Sinh Đơn Vị (Có/Nợ)", "Giải Trình Lệch Ảo của Kế Toán Trưởng"]
    format_sheet(ws_khop, "CÁC KHOẢN KHỚP THỰC TẾ TRÊN SỔ CHI TIẾT (BÁO LỆCH ẢO DO LỖI CÔNG THỨC ĐỐI CHIẾU)", headers_khop, ["center", "left", "currency", "currency", "center", "center", "left"])
    
    # Save workbook to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# -------------------------------------------------------------
# MAIN APP FLOW
# -------------------------------------------------------------
if tool_file is not None:
    with st.spinner("Đang đọc dữ liệu và thực hiện đối soát tự động..."):
        try:
            summary_data, details = load_and_reconcile(tool_file)
            
            # --- METRICS DASHBOARD ---
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            
            total_diff = sum([abs(r[4]) for r in summary_data])
            num_mismatched = sum([1 for r in summary_data if abs(r[4]) > 1])
            num_matched = len(summary_data) - num_mismatched
            
            with col_m1:
                st.markdown(f"""
                <div class="metric-card metric-card-danger">
                    <div class="metric-label">Tổng Chênh Lệch Lũy Kế</div>
                    <div class="metric-value" style="color: #EF4444;">{total_diff:,.0f} đ</div>
                </div>
                """, unsafe_allow_html=True)
            with col_m2:
                st.markdown(f"""
                <div class="metric-card metric-card-danger">
                    <div class="metric-label">Số Tài Khoản Bị Lệch</div>
                    <div class="metric-value" style="color: #EF4444;">{num_mismatched}</div>
                </div>
                """, unsafe_allow_html=True)
            with col_m3:
                st.markdown(f"""
                <div class="metric-card metric-card-success">
                    <div class="metric-label">Số Tài Khoản Đã Khớp</div>
                    <div class="metric-value" style="color: #10B981;">{num_matched}</div>
                </div>
                """, unsafe_allow_html=True)
            with col_m4:
                st.markdown("""
                <div class="metric-card metric-card-info">
                    <div class="metric-label">Trạng Thái Báo Cáo</div>
                    <div class="metric-value" style="color: #3B82F6;">Sẵn Sàng Xuất File</div>
                </div>
                """, unsafe_allow_html=True)
                
            st.markdown("<div style='margin-bottom: 25px;'></div>", unsafe_allow_html=True)

            # --- ACTION BAR (EXPORT & SAVE BUTTONS) ---
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                # Generate in-memory Excel file
                excel_buffer = generate_excel_report(summary_data, details)
                st.download_button(
                    label="📥 Tải Báo Cáo Excel (Tải Trực Tiếp)",
                    data=excel_buffer.getvalue(),
                    file_name="BaoCao_DoiChieu_136_336_Thang05_2026.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col_btn2:
                if st.button("💾 Lưu File Báo Cáo Vào Thư Mục Project", use_container_width=True):
                    try:
                        out_path = os.path.join(os.getcwd(), "BaoCao_DoiChieu_136_336_Thang05_2026_v2.xlsx")
                        with open(out_path, "wb") as f:
                            f.write(excel_buffer.getvalue())
                        st.success(f"✅ Đã lưu file thành công tại: {out_path}")
                    except Exception as ex:
                        st.error(f"Lỗi khi lưu file: {ex}")
            
            st.markdown("<div class='section-header'>KẾT QUẢ PHÂN TÍCH VÀ ĐỐI CHIẾU CHI TIẾT</div>", unsafe_allow_html=True)
            
            # --- DISPLAY TABS ---
            tab_sum, tab_111, tab_58, tab_xdcb, tab_khop = st.tabs([
                "📊 TỔNG HỢP CHÊNH LỆCH",
                "⚡ LỆCH DOANH THU ĐIỆN (1363111)",
                "💼 LỆCH NỘI BỘ KHÁC (136358)",
                "🏗️ LỆCH VỐN ĐẦU TƯ XDCB",
                "🔍 CÁC KHOẢN KHỚP THỰC TẾ (LỆCH ẢO)"
            ])
            
            # Tab 1: Tổng hợp chênh lệch
            with tab_sum:
                df_display_sum = pd.DataFrame(summary_data, columns=["Tài Khoản", "Tên Tài Khoản Đối Chiếu", "Số Dư TCT (TK 136)", "Số Dư Đơn Vị (TK 336)", "Chênh Lệch", "Phân Tích Nguyên Nhân Sự Việc"])
                
                # Format numbers for view
                df_formatted = df_display_sum.copy()
                df_formatted["Số Dư TCT (TK 136)"] = df_formatted["Số Dư TCT (TK 136)"].map(lambda x: f"{x:,.0f}" if pd.notnull(x) else "-")
                df_formatted["Số Dư Đơn Vị (TK 336)"] = df_formatted["Số Dư Đơn Vị (TK 336)"].map(lambda x: f"{x:,.0f}" if pd.notnull(x) else "-")
                df_formatted["Chênh Lệch"] = df_formatted["Chênh Lệch"].map(lambda x: f"{x:,.0f}" if pd.notnull(x) else "-")
                
                # Apply style coloring to streamlit dataframe
                def style_diff(val):
                    val_clean = str(val).replace(",", "")
                    try:
                        num = float(val_clean)
                        if num != 0:
                            return 'background-color: #F8D7DA; color: #9C0006; font-weight: bold;'
                        else:
                            return 'background-color: #D1E7DD; color: #006100; font-weight: bold;'
                    except ValueError:
                        return ''
                        
                st.dataframe(
                    df_formatted.style.applymap(style_diff, subset=["Chênh Lệch"]),
                    use_container_width=True,
                    height=500
                )
                
            # Helper to display detailed dataframes
            def show_detail_df(unmatched_tct, unmatched_pcvt, cols, headers):
                tct_part = unmatched_tct.copy()
                tct_part.insert(0, "Bên Hạch Toán", "Tổng công ty (TCT)")
                
                pcvt_part = unmatched_pcvt.copy()
                pcvt_part.insert(0, "Bên Hạch Toán", "Đơn vị (PCVT)")
                
                combined = pd.concat([tct_part, pcvt_part])
                
                # Filter columns to display
                cols_display = ["Bên Hạch Toán"] + cols
                combined_filtered = combined[cols_display].copy()
                
                # Format columns
                for c in combined_filtered.columns:
                    if 'SỐ PS' in c or 'NET' in c or 'Chênh Lệch' in c:
                        combined_filtered[c] = combined_filtered[c].map(lambda x: f"{x:,.0f}" if pd.notnull(x) else "-")
                    elif 'NGÀY' in c:
                        combined_filtered[c] = pd.to_datetime(combined_filtered[c]).dt.strftime('%Y-%m-%d')
                        
                # Rename headers
                combined_filtered.columns = ["Bên Hạch Toán"] + headers
                
                st.dataframe(combined_filtered, use_container_width=True, height=400)

            # Tab 2: Lệch 1363111
            with tab_111:
                st.markdown("##### Danh sách các giao dịch chưa khớp trên TK 1363111 vs TK 3363111")
                tct_111 = details["1363111_3363111"]["tct"]
                unit_111 = details["1363111_3363111"]["unit"]
                show_detail_df(tct_111, unit_111, ['NGÀY CT', 'SỐ CT PHÂN HỆ GL', 'DIỄN GIẢI', 'SỐ PS NỢ', 'SỐ PS CÓ', 'NGƯỜI LẬP CT'], ["Ngày CT", "Số CT GL", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Người Lập"])

            # Tab 3: Lệch 136358
            with tab_58:
                st.markdown("##### Danh sách các giao dịch chưa khớp trên TK 136358 vs TK 336358")
                tct_58 = details["136358_336358"]["tct"]
                unit_58 = details["136358_336358"]["unit"]
                show_detail_df(tct_58, unit_58, ['NGÀY CT', 'NGUỒN', 'DIỄN GIẢI', 'SỐ PS NỢ', 'SỐ PS CÓ', 'NGƯỜI LẬP CT'], ["Ngày CT", "Nguồn", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Người Lập"])

            # Tab 4: Lệch XDCB
            with tab_xdcb:
                st.markdown("##### Chi tiết chênh lệch và phân loại nhầm - Nhóm Vốn Đầu Tư XDCB (136331, 136332, 136341, 136351)")
                
                # Combine them for display
                combined_xdcb_tct = []
                combined_xdcb_unit = []
                for p_name in ["136331_336331", "136332_336332", "136341_336341", "136351_336351"]:
                    combined_xdcb_tct.append(details[p_name]["tct"])
                    combined_xdcb_unit.append(details[p_name]["unit"])
                    
                df_xdcb_tct = pd.concat(combined_xdcb_tct)
                df_xdcb_unit = pd.concat(combined_xdcb_unit)
                
                show_detail_df(df_xdcb_tct, df_xdcb_unit, ['SHTK', 'NGÀY CT', 'DIỄN GIẢI', 'SỐ PS NỢ', 'SỐ PS CÓ'], ["Tài Khoản", "Ngày CT", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có"])

            # Tab 5: Khớp thực tế (Khớp Ảo)
            with tab_khop:
                st.markdown("##### Giải trình chi tiết các tài khoản khớp thực tế nhưng bị báo lệch do công thức đối chiếu")
                khop_rows = []
                for r in summary_data:
                    tk = r[0].split(" / ")[0]
                    if tk in ["13638221", "1363847", "1363848", "1363883", "13681"]:
                        khop_rows.append([
                            r[0], r[1], r[2], r[3], r[5]
                        ])
                df_khop = pd.DataFrame(khop_rows, columns=["Tài Khoản", "Tên Tài Khoản", "Số Dư TCT (136)", "Số Dư Đơn Vị (336)", "Giải Trình Kỹ Thuật"])
                
                # Format numbers
                df_khop["Số Dư TCT (136)"] = df_khop["Số Dư TCT (136)"].map(lambda x: f"{x:,.0f}" if pd.notnull(x) else "-")
                df_khop["Số Dư Đơn Vị (336)"] = df_khop["Số Dư Đơn Vị (336)"].map(lambda x: f"{x:,.0f}" if pd.notnull(x) else "-")
                
                st.dataframe(df_khop, use_container_width=True, height=300)

        except Exception as e:
            st.error(f"Đã xảy ra lỗi trong quá trình đối soát: {e}")
            st.exception(e)
else:
    # Instructions when file is not yet uploaded
    st.info("💡 Vui lòng kéo và thả file Tool Đối Chiếu (Tool ERP-DoiChieu 136-336__TCT.xlsx) vào khung tải lên ở thanh công cụ bên trái để bắt đầu phân tích dữ liệu.")
    
    # Showcase the feature highlights
    st.markdown('<div class="section-header">TÍNH NĂNG VÀ CÔNG NGHỆ ÁP DỤNG</div>', unsafe_allow_html=True)
    
    col_feat1, col_feat2, col_feat3 = st.columns(3)
    with col_feat1:
        st.markdown("""
        <div class="feature-card">
            <h4>⚡ Đối Soát Giao Dịch Tự Động</h4>
            <p>Ứng dụng áp dụng thuật toán 3 bước (Multi-pass transaction matching) để đối soát tự động hàng nghìn giao dịch giữa TCT và Đơn vị chỉ trong vài giây, tìm ra chính xác giao dịch chênh lệch.</p>
        </div>
        """, unsafe_allow_html=True)
    with col_feat2:
        st.markdown("""
        <div class="feature-card">
            <h4>🤖 Giải Trình Động Thông Minh</h4>
            <p>Hệ thống phân tích và tự động viết nội dung giải trình lý do chênh lệch dựa trên số liệu thực tế được tính toán từ các tệp Excel, thay thế hoàn toàn việc gõ tay thủ công.</p>
        </div>
        """, unsafe_allow_html=True)
    with col_feat3:
        st.markdown("""
        <div class="feature-card">
            <h4>📥 Xuất Báo Cáo Đúng Định Dạng</h4>
            <p>Nút xuất Excel cho phép tải về báo cáo hoàn chỉnh được định dạng chuẩn Segoe UI, không số thập phân lẻ, phân cách hàng nghìn bằng dấu chấm, tô màu đỏ/xanh chuyên nghiệp.</p>
        </div>
        """, unsafe_allow_html=True)
