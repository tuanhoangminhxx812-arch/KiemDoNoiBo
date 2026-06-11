import os
import sys
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# Đường dẫn file
workspace_dir = r"c:\Users\Admin\OneDrive\PCVT\1. TCKT\Tổng hợp\KiemDoNoiBo"
source_file = os.path.join(workspace_dir, "Tool ERP-DoiChieu 136-336__TCT.xlsx")
output_file = os.path.join(workspace_dir, "BaoCao_DoiChieu_136_336_Thang05_2026.xlsx")

print("Reading data from source file...")
xl = pd.ExcelFile(source_file)
df_all = xl.parse("DATA_PHATSINH")

# Chuẩn hóa dữ liệu
df_all['SỐ PS NỢ'] = df_all['SỐ PS NỢ'].fillna(0).astype(float)
df_all['SỐ PS CÓ'] = df_all['SỐ PS CÓ'].fillna(0).astype(float)
df_all['SHTK'] = df_all['SHTK'].astype(str).str.split('.').str[0]
df_all['MÃ ĐVI'] = df_all['MÃ ĐVI'].astype(str).str.split('.').str[0]

# Phân chia TCT và PCVT
tct_all = df_all[df_all['MÃ ĐVI'] == '80100'].copy()
pcvt_all = df_all[df_all['MÃ ĐVI'] == '82900'].copy()

# Định nghĩa màu sắc (Professional Navy Theme)
NAVY_HEADER = "1B365D"       # Xanh navy đậm cho Header
NAVY_ACCENT = "E8EEF5"       # Zebra striping light blue
WHITE = "FFFFFF"
GRAY_TEXT = "595959"
BORDER_COLOR = "D9D9D9"      # Border mỏng
ALERT_YELLOW = "FFF3CD"      # Màu cảnh báo vàng nhạt
ALERT_GREEN = "D1E7DD"       # Màu khớp xanh nhạt
ALERT_RED = "F8D7DA"         # Màu chênh lệch đỏ nhạt

# Các hàm định dạng openpyxl
def apply_border(ws, start_row, end_row, start_col, end_col):
    thin = Side(border_style="thin", color=BORDER_COLOR)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

def format_ws(ws, title, headers, col_formats):
    # Set Title
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    ws.row_dimensions[1].height = 30
    
    # Write Headers at Row 3
    for idx, h in enumerate(headers):
        col = idx + 1
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28
    
    # Auto-fit columns and format number alignment
    for col_idx, fmt in enumerate(col_formats):
        col_letter = get_column_letter(col_idx + 1)
        # Apply style to data rows
        for r in range(4, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx + 1)
            cell.font = Font(name="Segoe UI", size=10)
            
            # Formatting
            if fmt == "currency":
                cell.number_format = '#,##0;[Red]-#,##0;"-";@'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "date":
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif fmt == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Zebra striping
            if r % 2 == 0:
                cell.fill = PatternFill(start_color=NAVY_ACCENT, end_color=NAVY_ACCENT, fill_type="solid")
                
        # Calculate width
        max_len = len(headers[col_idx])
        for r in range(4, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx + 1).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Set Borders
    apply_border(ws, 3, ws.max_row, 1, len(headers))
    ws.views.sheetView[0].showGridLines = True

# Tạo workbook mới
wb = openpyxl.Workbook()
# Remove default sheet
wb.remove(wb.active)

# -------------------------------------------------------------
# TAB 1: TONG_HOP_CHENH_LECH
# -------------------------------------------------------------
print("Đang tạo Sheet Tổng Hợp Chênh Lệch...")
ws_summary = wb.create_sheet(title="TONG_HOP_CHENH_LECH")

summary_data = [
    ["1363111 / 3363111", "Doanh thu tiền điện", -934199717138.0, 488183646170.0, -1422383363308.0, "TCT chưa hạch toán doanh thu điện chuyển lên từ đơn vị (1.42 nghìn tỷ). Lệch tk đối ứng khi chuyển tiền về TCT (TCT ghi 33688, đơn vị ghi 3363111). Lệch đối chiếu thu hộ qua cổng thanh toán."],
    ["1363112 / 3363112", "Doanh thu tiền điện vô công", -4548652836.0, -3876382947.0, -672269889.0, "PCVT kết chuyển doanh thu phản kháng tăng phải trả TCT (762.6 tr), TCT chưa ghi nhận. Lệch đối chiếu giải trừ thu hộ."],
    ["136314 / 336314", "Doanh thu QL dây, cáp", 7987850607.0, 8241423967.0, -253573360.0, "PCVT kết chuyển doanh thu QL dây cáp (253.5 tr), TCT chưa hạch toán ghi tăng phải thu tương ứng."],
    ["136331 / 336331", "Vốn ĐT XDCB", -392951167430.0, -399274041310.0, 6322873880.0, "PCVT kết chuyển giảm phải trả do quyết toán công trình hoàn thành về TCT (7.91 tỷ), TCT chưa hạch toán ghi Có. PCVT hạch toán tăng phải trả nhận điều động TSCĐ (2.53 tỷ) nhưng TCT chưa hạch toán đối ứng."],
    ["136332 / 336332", "Vốn ĐT XDCB khác", 0.0, 877629907.0, -877629907.0, "Lệch phân loại: PCVT hạch toán cấp vốn KHCB trên 336332 (Có), TCT hạch toán cấp vốn này vào 136331 (Nợ). Khớp 100% về tổng."],
    ["136341 / 336341", "Khấu hao TSCĐ nguồn vốn vay", -583672582158.0, -588783942615.0, 5111360457.0, "Lệch phân loại: PCVT kết chuyển khấu hao về TCT trên 336341, TCT hạch toán ghi nhận vào 136352. Khớp 100% chi tiết."],
    ["136351 / 336351", "Phải thu/trả Sửa chữa lớn", -17603701513.0, -18678057352.0, 1074355839.0, "PCVT kết chuyển chi phí SCL hoàn thành (1.07 tỷ) cấn trừ phải trả, TCT chưa hạch toán giảm phải thu tương ứng."],
    ["136358 / 336358", "Phải thu/trả nội bộ khác", 308607185793.0, 362318616683.0, -53711430890.0, "TCT chưa ghi nhận điện mua nội bộ đơn vị chuyển lên (Net=104.4 tỷ). Lệch đối chiếu cấp vốn lương/bảo hiểm (13.4 tỷ). Luân chuyển VTTB Net=0."],
    ["1363611 / 3363611", "Thuế GTGT ĐMT mái nhà", 4123490961.0, 2118534597.0, 2004956364.0, "TCT cấp vốn thuế GTGT đợt 1 (2.0B) ngày 20/05/2026, PCVT chưa ghi nhận nhận cấp vốn này trên 3363611."],
    ["1363613 / 3363613", "Thuế GTGT nội bộ khác", 5885031170.0, 5885031170.0, 0.0, "Khớp 100% cả số dư và phát sinh."],
    ["13638221 / 33638221", "LN vi phạm HĐ bán điện", 213770328.0, 213770328.0, 0.0, "Khớp 100%. File đối chiếu báo lệch ảo do công thức đối chiếu của TCT không cộng phát sinh trong tháng của TCT."],
    ["1363847 / 3363847", "LN hoạt động tài chính", 89994124.0, 89994124.0, 0.0, "Khớp 100%. File đối chiếu báo lệch ảo do công thức đối chiếu của TCT không cộng phát sinh trong tháng của TCT."],
    ["1363848 / 3363848", "LN thu nhập khác", -2638791984.0, -2638791984.0, 0.0, "Khớp 100%. File đối chiếu báo lệch ảo do công thức đối chiếu của TCT không cộng phát sinh trong tháng của TCT."],
    ["1363883 / 3363883", "Luân chuyển vật tư thiết bị", 16341791405.0, 16341791405.0, 0.0, "Khớp 100% về vật tư luân chuyển (17.7 tỷ). PCVT có hạch toán điều chỉnh cấn trừ nội bộ, TCT ghi nhận khớp ở tiểu khoản khác."],
    ["13681 / 33681", "Phải thu/trả nội bộ về chi phí", 15100381544.0, 15100381544.0, 0.0, "Phát sinh trong tháng khớp 100%. Chênh lệch số dư lũy kế đầu kỳ lịch sử (-717,044,052đ) từ các tháng trước để lại."]
]

# Write header on tab 1
headers_summary = ["Tài Khoản", "Tên Tài Khoản Đối Chiếu", "Số Dư TCT (TK 136)", "Số Dư Đơn Vị (TK 336)", "Chênh Lệch (TCT - Đơn vị)", "Phân Tích Nguyên Nhân Sự Việc"]
for r_idx, r in enumerate(summary_data):
    for c_idx, val in enumerate(r):
        ws_summary.cell(row=r_idx + 4, column=c_idx + 1, value=val)

# Formats for columns
format_summary = ["center", "left", "currency", "currency", "currency", "left"]
format_ws(ws_summary, "BẢNG TỔNG HỢP CHÊNH LỆCH CÔNG NỢ NỘI BỘ TK 136/336 - THÁNG 05/2026", headers_summary, format_summary)

# Highlight discrepancy rows
fill_green = PatternFill(start_color=ALERT_GREEN, end_color=ALERT_GREEN, fill_type="solid")
fill_red = PatternFill(start_color=ALERT_RED, end_color=ALERT_RED, fill_type="solid")

for r_idx in range(4, ws_summary.max_row + 1):
    diff = ws_summary.cell(row=r_idx, column=5).value
    # Highlight discrepancy cell
    cell_diff = ws_summary.cell(row=r_idx, column=5)
    if diff != 0:
        cell_diff.fill = fill_red
        cell_diff.font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
    else:
        cell_diff.fill = fill_green
        cell_diff.font = Font(name="Segoe UI", size=10, bold=True, color="006100")

# -------------------------------------------------------------
# TAB 2: LECH_1363111_3363111
# -------------------------------------------------------------
print("Đang tạo Sheet Chi Tiết Lệch TK 1363111...")
ws_111 = wb.create_sheet(title="LECH_1363111_3363111")

headers_111 = ["Bên Hạch Toán", "Ngày CT", "Số CT GL", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Người Lập", "Phân Tích Nghiệp Vụ Chênh Lệch"]

data_111 = [
    ["Đơn vị (PCVT)", "2026-05-31", 1622, "K/c 51111 > 3363111 (k/c doanh thu điện)", 0, 1420763729435.0, "DUYENTT", "Đơn vị hạch toán kết chuyển doanh thu điện Vũng Tàu chuyển lên TCT. TCT chưa hạch toán ghi nhận khoản doanh thu này (làm lệch -1.42 nghìn tỷ)."],
    ["TCT", "2026-05-31", 5092, "KC PTNB TIEN DIEN THANG 5-2026 - PCVT", 0, 1006728865803.0, "ANHTTL", "TCT hạch toán ghi giảm phải thu điện của đơn vị (bút toán bù trừ thu tiền). Khớp với số tiền điện thực tế đơn vị nộp về."],
    ["TCT", "2026-05-31", 5045, "KC bù trừ thuế GTGT điện nọi bộ với thuế GTGT bán điện T4.2026", 103307926903.0, 0, "MINHDQ", "TCT ghi tăng phải thu thuế GTGT điện nội bộ. Khớp 100% với đơn vị hạch toán."],
    ["Đơn vị (PCVT)", "2026-05-31", 1392, "HT bù trừ tiền thuế GTGT T04/2026", 0, 103307926903.0, "DUYENTT", "Đơn vị hạch toán tăng phải trả thuế GTGT điện nội bộ. Khớp 100% với TCT."],
    ["Đơn vị (PCVT)", "Chuyển tiền", "35 dòng", "CTĐLVT chuyển tiền điện về TCT ĐL HCM qua các NH", 1006630413068.0, 0, "CHAULTN", "Đơn vị ghi Nợ giảm phải trả điện. TCT lại hạch toán các khoản thu này trên TK 33688 (Có) chứ không ghi nhận trực tiếp vào TK 1363111, làm lệch tài khoản đối ứng."],
    ["TCT", "Thu tiền", "36 dòng", "Nhập thu tiền bán điện - PCVT (trên TK 33688)", 0, 1007030102392.0, "VIETDT", "TCT nhận tiền đơn vị chuyển về hạch toán trên TK 33688. Số liệu chênh lệch 399,689,324đ so với đơn vị do kiểm đếm lệch thời điểm."],
    ["Đơn vị (PCVT)", "Giải trừ", "27 dòng", "Giải trừ tiền điện thu hộ CTĐLVT qua các cổng Airpay, Momo...", 458342163050.0, 0, "QUYENNN", "Đơn vị hạch toán giảm phải trả điện qua các cổng thanh toán. TCT mới chỉ hạch toán ghi Có 1363111 số tiền 107,450,427,113đ, làm lệch 350.8 tỷ."],
    ["TCT", "Cổng thanh toán", "205 dòng", "Thu qua các cổng thanh toán Airpay, Momo, Payoo...", 0, 107450427113.0, "VIETDT", "TCT ghi nhận thu qua các cổng thanh toán. Cần đối chiếu lại toàn bộ để hạch toán bổ sung phần đơn vị đã giải trừ (350.8 tỷ)."]
]

for r_idx, r in enumerate(data_111):
    for c_idx, val in enumerate(r):
        ws_111.cell(row=r_idx + 4, column=c_idx + 1, value=val)

format_111 = ["center", "center", "center", "left", "currency", "currency", "center", "left"]
format_ws(ws_111, "CHI TIẾT CÁC GIAO DỊCH CHÊNH LỆCH LỚN - TK 1363111 vs TK 3363111", headers_111, format_111)

# -------------------------------------------------------------
# TAB 3: LECH_136358_336358
# -------------------------------------------------------------
print("Đang tạo Sheet Chi Tiết Lệch TK 136358...")
ws_58 = wb.create_sheet(title="LECH_136358_336358")

headers_58 = ["Bên Hạch Toán", "Ngày CT", "Nguồn", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Người Lập", "Phân Tích Nghiệp Vụ Chênh Lệch"]

data_58 = [
    ["Đơn vị (PCVT)", "2026-05-31", "GL-MANUAL", "HT CP điện mua nội bộ T5/2026 (HĐ 819)", 0, 1409965472398.0, "THANH5DT", "Đơn vị hạch toán tăng phải trả điện mua nội bộ từ TCT. TCT chưa hạch toán ghi nhận trên 136358."],
    ["Đơn vị (PCVT)", "2026-05-31", "GL-MANUAL", "HT bù trừ chi phí điện mua nội bộ trong tháng", 1305523222043.0, 0, "THANH5DT", "Đơn vị hạch toán bù trừ giảm phải trả điện. TCT chưa hạch toán ghi nhận (làm chênh lệch ròng 104.4 tỷ)."],
    ["TCT", "Trong tháng", "AR-Misc", "Cấp vốn lương, bảo hiểm, điện mặt trời cho đơn vị", 35554965957.0, 0, "TAMLTM", "TCT ghi Nợ phải thu khác về cấp vốn. Đơn vị mới hạch toán ghi Có nhận cấp vốn là 22,119,626,146đ. Chênh lệch 13.4 tỷ cần đối chiếu lại."],
    ["Đơn vị (PCVT)", "Trong tháng", "AR-Misc", "Nhập tiền cấp lương, bảo hiểm, ĐMT từ TCT", 0, 22119626146.0, "THAOPTT", "Đơn vị ghi nhận nhận tiền cấp vốn từ TCT. Có sự lệch pha 13.4 tỷ so với số TCT cấp."],
    ["Đơn vị (PCVT)", "2026-05-31", "GL-MANUAL", "Kết chuyển công nợ nội bộ tháng 05/2026 (ghi âm)", -47447511170.0, 0, "DUYENTT", "Đơn vị hạch toán kết chuyển giảm công nợ nội bộ. TCT chưa hạch toán ghi nhận."],
    ["Đơn vị (PCVT)", "2026-05-31", "GL-MANUAL", "PCVT kết chuyển lãi vay TK635111 về Tcty", 3758552437.0, 0, "DUYENTT", "Đơn vị hạch toán kết chuyển lãi vay giảm phải trả nội bộ. TCT chưa ghi nhận đối ứng."],
    ["Đơn vị (PCVT)", "Trong tháng", "Inventory", "Nhập/Xuất vật tư thiết bị đi thí nghiệm kiểm định", 1619307789.0, 2977343277.0, "THANH5DT", "Giao dịch nội bộ của đơn vị mang vật tư đi kiểm định (Net = 0), TCT không ghi nhận hoặc ghi nhận ở tài khoản khác."]
]

for r_idx, r in enumerate(data_58):
    for c_idx, val in enumerate(r):
        ws_58.cell(row=r_idx + 4, column=c_idx + 1, value=val)

format_58 = ["center", "center", "center", "left", "currency", "currency", "center", "left"]
format_ws(ws_58, "CHI TIẾT CÁC GIAO DỊCH CHÊNH LỆCH LỚN - TK 136358 vs TK 336358", headers_58, format_58)

# -------------------------------------------------------------
# TAB 4: LECH_XDCB
# -------------------------------------------------------------
print("Đang tạo Sheet Chi Tiết Lệch Vốn Đầu Tư XDCB...")
ws_xdcb = wb.create_sheet(title="LECH_XDCB")

headers_xdcb = ["Tài Khoản", "Bên Hạch Toán", "Ngày CT", "Diễn Giải", "Phát Sinh Nợ", "Phát Sinh Có", "Chênh Lệch", "Nguyên Nhân & Phương Án Điều Chỉnh"]

data_xdcb = [
    ["136331 / 336331", "Đơn vị (PCVT)", "2026-05-31", "Kết chuyển công trình quyết toán hoàn thành (CT23018, CT23009)", 7911429900.0, 0, 7911429900.0, "PCVT kết chuyển ghi giảm phải trả do công trình hoàn thành bàn giao cho TCT. TCT chưa hạch toán ghi Có giảm phải thu tương ứng. TCT cần ghi Có 136331 số tiền này."],
    ["136331 / 336331", "Đơn vị (PCVT)", "2026-05-31", "Nhận điều động tài sản cố định nội bộ TCT", 0, 2534843522.0, -2534843522.0, "PCVT ghi tăng phải trả nhận TSCĐ bàn giao từ TCT. TCT chưa ghi nhận giảm tài sản/tăng phải thu tương ứng. TCT cần hạch toán ghi nhận."],
    ["136331 / 336331", "TCT / PCVT", "Trong tháng", "Cấp vốn đầu tư XDCB thanh toán gốc, lãi vay", 14964537632.0, 15282854491.0, -318316859.0, "Chênh lệch số liệu cấp vốn đầu tư giữa TCT (ghi Nợ) và PCVT (ghi Có). Cần rà soát đối chiếu chi tiết chứng từ."],
    ["136332 / 336332", "TCT / PCVT", "2026-05-31", "Cấp vốn KHCB khác thanh toán CP kiểm toán ATAX, Phú Tài Phát...", 0, 877629907.0, -877629907.0, "Lệch phân loại: TCT hạch toán cấp vốn vào TK 136331 (Nợ), nhưng PCVT hạch toán nhận vào TK 336332 (Có). Tổng số tiền khớp nhau. TCT cần điều chuyển phân loại từ 136331 sang 136332."],
    ["136341 / 336341", "TCT / PCVT", "2026-05-31", "Kết chuyển chi phí khấu hao TSCĐ nguồn vốn vay", 0, 5111360457.0, -5111360457.0, "Lệch phân loại: PCVT ghi giảm phải trả khấu hao trên 336341, TCT hạch toán ghi nhận nhưng đưa nhầm vào TK 136352. Thực tế khớp 100%. TCT cần chuyển phân loại từ 136352 sang 136341."],
    ["136351 / 336351", "Đơn vị (PCVT)", "2026-05-31", "Kết chuyển chi phí sửa chữa lớn hoàn thành cấn trừ công nợ", 1074355839.0, 0, 1074355839.0, "PCVT kết chuyển giảm phải trả chi phí SCL hoàn thành bàn giao cho TCT. TCT chưa hạch toán ghi Có giảm phải thu tương ứng. TCT cần hạch toán bổ sung."]
]

for r_idx, r in enumerate(data_xdcb):
    for c_idx, val in enumerate(r):
        ws_xdcb.cell(row=r_idx + 4, column=c_idx + 1, value=val)

format_xdcb = ["center", "center", "center", "left", "currency", "currency", "currency", "left"]
format_ws(ws_xdcb, "CHI TIẾT CHÊNH LỆCH VÀ PHÂN LOẠI NHẦM - NHÓM VỐN ĐẦU TƯ XDCB & KHẤU HAO", headers_xdcb, format_xdcb)

# -------------------------------------------------------------
# TAB 5: KHOAN_KHOP_AO
# -------------------------------------------------------------
print("Đang tạo Sheet Các Khoản Khớp Thực Tế...")
ws_khop = wb.create_sheet(title="KHOAN_KHOP_AO")

headers_khop = ["Tài Khoản", "Tên Tài Khoản", "Số Dư TCT", "Số Dư Đơn Vị", "Phát Sinh TCT (Nợ/Có)", "Phát Sinh Đơn Vị (Có/Nợ)", "Giải Trình Lệch Ảo của Kế Toán Trưởng"]

data_khop = [
    ["13638221 / 33638221", "LN vi phạm HĐ bán điện", 213770328.0, 213770328.0, 2468617.0, 2468617.0, "Khớp 100%. Bảng đối chiếu báo lệch do công thức ở TCT chỉ lấy Số dư đầu kỳ của TCT (211,301,711) so với Số dư cuối kỳ của Đơn vị (213,770,328). Không có chênh lệch thực tế."],
    ["1363847 / 3363847", "LN hoạt động tài chính", 89994124.0, 89994124.0, 2095232.0, 2095232.0, "Khớp 100%. Bảng đối chiếu báo lệch do công thức ở TCT chỉ lấy Số dư đầu kỳ của TCT (87,898,892) so với Số dư cuối kỳ của Đơn vị (89,994,124). Không có chênh lệch thực tế."],
    ["1363848 / 3363848", "LN thu nhập khác", -2638791984.0, -2638791984.0, 8999937.0, 8999937.0, "Khớp 100%. Bảng đối chiếu báo lệch do công thức ở TCT chỉ lấy Số dư đầu kỳ của TCT (-2,647,791,921) so với Số dư cuối kỳ của Đơn vị (-2,638,791,984). Không có chênh lệch thực tế."],
    ["1363883 / 3363883", "Luân chuyển vật tư thiết bị", 16341791405.0, 16341791405.0, 17718778000.0, 17718778000.0, "Khớp 100% về giá trị vật tư điều động trong tháng (17.7 tỷ). Phần chênh lệch số phát sinh là do đơn vị hạch toán cấn trừ công nợ nội bộ cấn trừ vốn, TCT hạch toán khớp ở các tk chi tiết khác."],
    ["13681 / 33681", "Phải thu/trả khác về chi phí", 15100381544.0, 15100381544.0, "Nợ: 452.7 tr / Có: 590.8 tr", "Nợ: 590.8 tr / Có: 452.7 tr", "Phát sinh trong tháng khớp 100%. Chênh lệch số dư cuối kỳ thực chất là lệch lũy kế đầu kỳ lịch sử để lại từ các tháng trước (-717,044,052đ). Trong tháng không phát sinh thêm lệch."]
]

for r_idx, r in enumerate(data_khop):
    for c_idx, val in enumerate(r):
        ws_khop.cell(row=r_idx + 4, column=c_idx + 1, value=val)

format_khop = ["center", "left", "currency", "currency", "center", "center", "left"]
format_ws(ws_khop, "CÁC KHOẢN KHỚP THỰC TẾ TRÊN SỔ CHI TIẾT (BÁO LỆCH ẢO DO LỖI CÔNG THỨC ĐỐI CHIẾU)", headers_khop, format_khop)

# Lưu workbook
print("Đang lưu workbook...")
try:
    wb.save(output_file)
    print(f"Đã xuất báo cáo đối chiếu công nợ nội bộ thành công ra file: {output_file}")
except PermissionError:
    base, ext = os.path.splitext(output_file)
    output_file_v2 = f"{base}_v2{ext}"
    wb.save(output_file_v2)
    print(f"Lưu ý: File gốc đang mở, báo cáo đã được lưu tạm ra file mới: {output_file_v2}")
